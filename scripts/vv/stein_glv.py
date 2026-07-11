"""Stein et al. 2013 → :class:`GLVDataset` adapter (the deferred real-data ingestion).

Stein, Bucci, Toussaint, ... Xavier (2013) *PLoS Comput Biol* 9(12): e1003388,
"Ecological Modeling from Time-Series Inference: Insight into Dynamics and Stability of
Intestinal Microbiota" (CC-BY). Dataset S1 is a single ``.xlsx``:

* sheet ``Y`` — the processed taxa densities (11 groups × up to 11 timepoints), grouped by
  ``Population`` (experimental condition) × ``Replicate`` (colony), plus a **Clindamycin
  signal** row = the antibiotic pulse ``u(t)`` (a 1-day unit pulse at ``t = 0``).
* sheets ``F`` / ``D_lambda`` / ``MmuE`` — the authors' OWN fitted growth / interaction /
  susceptibility parameters (``MmuE`` = paper Fig. 2, genera sorted by susceptibility ε).
  We read those only as an independent **ground-truth reference** to score NUDGE against —
  NUDGE itself never sees them.

**The experimental design (verified from the file):** three conditions, three colonies each.

* **Population 1** — NO clindamycin (4 timepoints t∈{0,2,6,13}); the undisturbed community.
* **Population 2** — clindamycin pulse, C. difficile does NOT bloom (stays 0).
* **Population 3** — clindamycin pulse + C. difficile CHALLENGE → C. difficile blooms to
  ≈2–3 by t≈12–16 (competitive release after commensals are killed).

**The NUDGE contrast.** reference = Population 1 (no drug ≡ insusceptible, ε=0 under the
pulse), perturbed = Population 3 (drug on). Both share the clindamycin pulse ``u(t)``; the
reference is fit with ε≈0 (it shows no crash), so a taxon that crashes ONLY in the perturbed
group during the pulse window is the identifiable ``ε`` (direct-kill) signature — while
C. difficile's LATE, interaction-mediated bloom is NOT pulse-locked, so NUDGE must abstain.

This module is a **new, additive real-data adapter** (``scripts/vv/``). It imports the SHIPPED
:mod:`nudge.inference.lotka_volterra` read-only and does not modify it, ``fit.py``, or
``core/``.
"""

from __future__ import annotations

import warnings
from dataclasses import dataclass
from pathlib import Path

import numpy as np

from nudge.inference.lotka_volterra import GLVDataset, GLVParams

# The 11 groups, in the sheet-Y row order (rows 6..16).
STEIN_TAXA: tuple[str, ...] = (
    "undefined_genus_of_Enterobacteriaceae",  # 0
    "Blautia",  # 1
    "Barnesiella",  # 2
    "undefined_genus_of_unclassified_Mollicutes",  # 3
    "undefined_genus_of_Lachnospiraceae",  # 4
    "Akkermansia",  # 5
    "Clostridium_difficile",  # 6  <- the pathogen
    "unclassified_Lachnospiraceae",  # 7
    "Coprobacillus",  # 8
    "Enterococcus",  # 9
    "Other",  # 10
)
CDIFF_IDX = 6

#: Authors' published per-taxon antibiotic susceptibility ε (sheet ``MmuE``, Fig. 2).
#: Strongly negative = directly killed by clindamycin; positive = directly promoted.
#: C. difficile's ε≈−0.31 is SMALL — its bloom is interaction-mediated, not direct kill.
PUBLISHED_EPS: dict[str, float] = {
    "Barnesiella": -3.2926,
    "undefined_genus_of_Lachnospiraceae": -3.0354,
    "unclassified_Lachnospiraceae": -2.0909,
    "Other": -1.9395,
    "Blautia": -1.3491,
    "undefined_genus_of_unclassified_Mollicutes": -1.1018,
    "Akkermansia": -0.9245,
    "Coprobacillus": -0.7940,
    "Clostridium_difficile": -0.3127,
    "Enterococcus": 1.0671,
    "undefined_genus_of_Enterobacteriaceae": 3.7009,
}
#: Authors' published per-taxon intrinsic growth α (sheet ``MmuE`` "Growth" column).
PUBLISHED_GROWTH: dict[str, float] = {
    "Barnesiella": 0.3681,
    "undefined_genus_of_Lachnospiraceae": 0.3102,
    "unclassified_Lachnospiraceae": 0.3561,
    "Other": 0.5401,
    "Blautia": 0.7090,
    "undefined_genus_of_unclassified_Mollicutes": 0.4706,
    "Akkermansia": 0.2297,
    "Coprobacillus": 0.8300,
    "Clostridium_difficile": 0.3918,
    "Enterococcus": 0.2908,
    "undefined_genus_of_Enterobacteriaceae": 0.3237,
}


@dataclass(frozen=True)
class SteinRaw:
    """The parsed sheet-Y contents: per (population, replicate) trajectory ensembles."""

    #: keyed by (population, replicate) -> (T,) times and (T, 11) taxa-density matrix.
    times: dict[tuple[int, int], np.ndarray]
    dens: dict[tuple[int, int], np.ndarray]
    clinda: dict[tuple[int, int], np.ndarray]

    def replicates(self, population: int) -> list[tuple[int, int]]:
        keys = [k for k in self.times if k[0] == population]
        return sorted(keys)


def load_stein(path: str | Path) -> SteinRaw:
    """Parse Dataset S1 sheet ``Y`` into per-(population, replicate) density trajectories."""
    import openpyxl

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Y"]
    ncol = ws.max_column

    def rowvals(r: int) -> list[object]:
        return [ws.cell(r, c).value for c in range(2, ncol + 1)]

    pop = rowvals(1)
    rep = rowvals(2)
    trow = rowvals(4)
    clinda_row = rowvals(18)
    # taxa densities live on rows 6..16 (11 taxa), one column per sample.
    dens_rows = np.array(
        [[ws.cell(r, c).value for c in range(2, ncol + 1)] for r in range(6, 17)],
        dtype=float,
    )  # (11, n_samples)

    times: dict[tuple[int, int], list[float]] = {}
    cols: dict[tuple[int, int], list[int]] = {}
    clinda: dict[tuple[int, int], list[float]] = {}
    for j, (p, rp) in enumerate(zip(pop, rep, strict=True)):
        key = (int(p), int(rp))
        times.setdefault(key, []).append(float(trow[j]))
        cols.setdefault(key, []).append(j)
        clinda.setdefault(key, []).append(float(clinda_row[j]))

    out_t: dict[tuple[int, int], np.ndarray] = {}
    out_d: dict[tuple[int, int], np.ndarray] = {}
    out_c: dict[tuple[int, int], np.ndarray] = {}
    for key, ts in times.items():
        order = np.argsort(ts)
        idx = np.array(cols[key])[order]
        out_t[key] = np.array(ts)[order]
        out_d[key] = dens_rows[:, idx].T  # (T, 11)
        out_c[key] = np.array(clinda[key])[order]
    return SteinRaw(times=out_t, dens=out_d, clinda=out_c)


# --------------------------------------------------------------------------- #
# taxon aggregation to k functional groups (by published susceptibility tier)
# --------------------------------------------------------------------------- #
# C. difficile is ALWAYS its own group (index-fixed as the last group) so it can be the
# attribution target at every k. The remaining 10 taxa are merged by ε-susceptibility tier
# (a functional / taxonomic grouping), coarsest at k=2, fully split at k=11.
_STRONG_SUPP = [2, 4, 7, 10]  # ε < -1.5  (Barnesiella, Lachno×2, Other)
_MOD_SUPP = [1, 3, 5, 8]  # -1.5 < ε < 0 (Blautia, Mollicutes, Akkermansia, Coprobacillus)
_PROMOTED = [0, 9]  # ε > 0 (Enterobacteriaceae, Enterococcus)


def aggregation(k: int) -> tuple[list[list[int]], list[str]]:
    """Return (group member-lists, group labels) for a k-group aggregation.

    C. difficile is always the final singleton group. Grouping is by published
    susceptibility tier — the natural functional axis for an antibiotic contrast.
    """
    others = [i for i in range(11) if i != CDIFF_IDX]
    if k == 2:
        groups = [others]
        labels = ["background"]
    elif k == 3:
        groups = [_STRONG_SUPP + _MOD_SUPP, _PROMOTED]
        labels = ["suppressed", "promoted"]
    elif k == 5:
        groups = [_STRONG_SUPP, _MOD_SUPP, [0], [9]]
        labels = ["strong_suppressed", "mod_suppressed", "Enterobacteriaceae", "Enterococcus"]
    elif k == 8:
        groups = [[2], [4], [7, 10], [1, 3], [5, 8], [9], [0]]
        labels = [
            "Barnesiella", "Lachnospiraceae_u", "Lachno_other", "Blautia_Mollicutes",
            "Akkermansia_Coprobacillus", "Enterococcus", "Enterobacteriaceae",
        ]
    elif k == 11:
        groups = [[i] for i in others]
        labels = [STEIN_TAXA[i] for i in others]
    else:
        raise ValueError(f"unsupported k={k} (use 2,3,5,8,11)")
    groups.append([CDIFF_IDX])
    labels.append("Clostridium_difficile")
    assert sorted(i for g in groups for i in g) == list(range(11))
    return groups, labels


def _interp_traj(times: np.ndarray, dens: np.ndarray, grid: np.ndarray) -> np.ndarray:
    """Linearly interpolate a (T, S) trajectory onto ``grid`` (abundances, clipped ≥ 0)."""
    out = np.empty((len(grid), dens.shape[1]))
    for s in range(dens.shape[1]):
        out[:, s] = np.interp(grid, times, dens[:, s])
    return np.clip(out, 0.0, None)


def _aggregate(dens: np.ndarray, groups: list[list[int]]) -> np.ndarray:
    """Sum member densities within each group → (T, k). Densities are additive."""
    return np.stack([dens[:, g].sum(axis=1) for g in groups], axis=1)


def build_stein_dataset(
    raw: SteinRaw,
    *,
    k: int,
    reference_pop: int = 1,
    perturbed_pop: int = 3,
    t_grid: np.ndarray | None = None,
    dt: float = 0.05,
    pulse_len: float = 1.0,
    normalize: bool = True,
    baseline: GLVParams | None = None,
) -> tuple[GLVDataset, list[str], int]:
    """Assemble a :class:`GLVDataset` from Stein Population ``reference_pop`` vs ``perturbed_pop``.

    Aggregates the 11 taxa to ``k`` groups, interpolates every colony's trajectory onto a
    shared observation grid, and applies the clindamycin pulse ``u(t)=1`` on ``[0, pulse_len)``
    as the external input. C. difficile is the final group.

    ``normalize`` rescales each group's abundance to O(1) (per-group 95th-percentile over
    reference∪perturbed, floored). This is **α/ε-invariant**: gLV is scale-covariant in ``x``
    (``xᵢ→xᵢ/sᵢ`` only rescales ``βᵢⱼ→βᵢⱼsⱼ``, leaving intrinsic growth ``α`` and susceptibility
    ``ε`` — the attribution axes — unchanged), and it keeps the RK4 integrator out of the stiff
    blow-up regime that raw metagenomic densities (spanning ~1e-5..12) drive it into. Returns
    ``(dataset, group_labels, cdiff_group_index)``.
    """
    groups, labels = aggregation(k)
    cdiff_group = len(groups) - 1
    if t_grid is None:
        # perturbed native times, truncated to the reference's observed range (no extrapolation).
        ref_max = max(raw.times[key].max() for key in raw.replicates(reference_pop))
        pert_t = raw.times[raw.replicates(perturbed_pop)[0]]
        t_grid = pert_t[pert_t <= ref_max + 1e-9]
    t_grid = np.asarray(t_grid, dtype=float)

    def ensemble(pop: int) -> np.ndarray:
        trajs = []
        for key in raw.replicates(pop):
            interp = _interp_traj(raw.times[key], raw.dens[key], t_grid)
            trajs.append(_aggregate(interp, groups))
        return np.stack(trajs).astype(np.float32)  # (R, T, k)

    reference = ensemble(reference_pop)
    perturbed = ensemble(perturbed_pop)

    scale = np.ones(len(groups), dtype=np.float32)
    if normalize:
        both = np.concatenate([reference, perturbed], axis=0)  # (2R, T, k)
        scale = np.maximum(np.percentile(both, 95, axis=(0, 1)), 0.1).astype(np.float32)
        reference = reference / scale
        perturbed = perturbed / scale

    t_max = float(t_grid.max())
    n_steps = int(round(t_max / dt))
    grid_t = np.arange(n_steps + 1) * dt
    obs_idx = np.clip(np.round(t_grid / dt).astype(int), 0, n_steps)
    u_grid = ((grid_t[:-1] >= 0.0) & (grid_t[:-1] < pulse_len)).astype(np.float32)

    if baseline is None:
        # a placeholder; attribute_glv re-fits the baseline from the reference replicates.
        n = len(groups)
        baseline = GLVParams(
            alpha=np.full(n, 0.4), beta=-np.eye(n), eps=np.zeros(n)
        )

    ground_truth = {
        "mechanism": "real-data (unknown)",
        "target": cdiff_group,
        "reference_pop": reference_pop,
        "perturbed_pop": perturbed_pop,
        "k": k,
        "labels": labels,
        "scale": scale.tolist(),
    }
    ds = GLVDataset(
        reference=reference,
        perturbed=perturbed,
        t_obs=t_grid,
        u_grid=u_grid,
        obs_idx=obs_idx.astype(np.int64),
        dt=dt,
        baseline=baseline,
        ground_truth=ground_truth,
    )
    return ds, labels, cdiff_group


def published_group_eps(k: int) -> list[float]:
    """Abundance-unweighted mean published ε over each aggregated group (for scoring only)."""
    groups, _ = aggregation(k)
    return [float(np.mean([PUBLISHED_EPS[STEIN_TAXA[i]] for i in g])) for g in groups]
